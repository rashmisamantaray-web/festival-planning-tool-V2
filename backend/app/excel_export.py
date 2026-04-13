"""
Excel export – generates a formatted multi-sheet workbook.

One sheet per level. Editable cells highlighted in yellow.
Now uses string keys ("current", "ref1", etc.) instead of integer years.
"""

from __future__ import annotations

from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, size=10)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
PCT_FORMAT = "0.00%"
NUM_FORMAT = "#,##0.00"
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, row: int, col_start: int, col_end: int):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def _key_label(key: str) -> str:
    """Convert key like 'ref1' to a display label like 'Ref 1'."""
    if key == "current":
        return "Current"
    return key.replace("ref", "Ref ")


def _write_city_sheet(ws, city_data: dict):
    hist_keys = city_data.get("historical_keys", [])
    all_keys = city_data.get("all_keys", [])
    records = city_data["data"]

    headers = ["City"]
    for k in all_keys:
        lbl = _key_label(k)
        headers += [f"Wk {lbl}", f"Day {lbl}", f"Baseline {lbl}", f"Actual {lbl}"]
    for k in hist_keys:
        headers.append(f"Pristine Drop {_key_label(k)}")
    for k in hist_keys:
        headers.append(f"Base Corr Drop {_key_label(k)}")
    headers += ["Override Row 1", "Override Row 2", "Final Impact %"]

    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, 1, 1, len(headers))

    for ri, rec in enumerate(records, 2):
        col = 1
        ws.cell(row=ri, column=col, value=rec["city_name"]); col += 1

        for k in all_keys:
            yd = rec["years"].get(k, {})
            ws.cell(row=ri, column=col, value=yd.get("week", "")); col += 1
            ws.cell(row=ri, column=col, value=yd.get("day_name", "")); col += 1
            ws.cell(row=ri, column=col, value=yd.get("baseline", 0)).number_format = NUM_FORMAT; col += 1
            ws.cell(row=ri, column=col, value=yd.get("actual", 0)).number_format = NUM_FORMAT; col += 1

        for k in hist_keys:
            val = rec["years"].get(k, {}).get("pristine_drop_pct", 0) / 100
            ws.cell(row=ri, column=col, value=val).number_format = PCT_FORMAT; col += 1

        for k in hist_keys:
            val = rec["years"].get(k, {}).get("base_corrected_drop_pct", 0) / 100
            ws.cell(row=ri, column=col, value=val).number_format = PCT_FORMAT; col += 1

        c1 = ws.cell(row=ri, column=col, value=rec.get("override_row1", 0) / 100)
        c1.number_format = PCT_FORMAT; c1.fill = YELLOW_FILL; col += 1

        c2 = ws.cell(row=ri, column=col, value=rec.get("override_row2", 0) / 100)
        c2.number_format = PCT_FORMAT; c2.fill = YELLOW_FILL; col += 1

        final_cell = ws.cell(row=ri, column=col, value=rec.get("final_impact_pct", 0) / 100)
        final_cell.number_format = PCT_FORMAT
        final_cell.font = Font(bold=True)

    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14


def _write_indexed_sheet(ws, level_data: dict, group_fields: list[str], parent_drop_label: str):
    hist_keys = level_data.get("historical_keys", [])
    records = level_data["data"]
    cur_key = level_data.get("current_key", "current")

    headers = [f.replace("_", " ").title() for f in group_fields]
    for k in hist_keys:
        lbl = _key_label(k)
        headers += [f"Base {lbl}", f"Fest {lbl}", f"Drop {lbl}"]
    for k in hist_keys:
        headers.append(f"Base Corr {_key_label(k)}")
    headers += [
        f"Base {_key_label(cur_key)}",
        "Final %",
        "Drop With Current %",
        parent_drop_label,
        "Final After Indexing %",
    ]

    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, 1, 1, len(headers))

    for ri, rec in enumerate(records, 2):
        col = 1
        for f in group_fields:
            ws.cell(row=ri, column=col, value=rec.get(f, "")); col += 1

        for k in hist_keys:
            yd = rec.get("years", {}).get(k, {})
            ws.cell(row=ri, column=col, value=yd.get("baseline", 0)).number_format = NUM_FORMAT; col += 1
            ws.cell(row=ri, column=col, value=yd.get("actual", 0)).number_format = NUM_FORMAT; col += 1
            val = yd.get("pristine_drop_pct", 0) / 100
            ws.cell(row=ri, column=col, value=val).number_format = PCT_FORMAT; col += 1

        for k in hist_keys:
            val = rec.get("years", {}).get(k, {}).get("base_corrected_drop_pct", 0) / 100
            ws.cell(row=ri, column=col, value=val).number_format = PCT_FORMAT; col += 1

        cur_bl = rec.get("years", {}).get(cur_key, {}).get("baseline", 0)
        ws.cell(row=ri, column=col, value=cur_bl).number_format = NUM_FORMAT; col += 1

        final_cell = ws.cell(row=ri, column=col, value=rec.get("final_pct", 0) / 100)
        final_cell.number_format = PCT_FORMAT
        final_cell.fill = YELLOW_FILL
        col += 1

        ws.cell(row=ri, column=col, value=rec.get("drop_with_current_pct", 0) / 100).number_format = PCT_FORMAT; col += 1

        parent_val = (rec.get("city_drop_pct", 0) or rec.get("subcat_drop_pct", 0)) / 100
        ws.cell(row=ri, column=col, value=parent_val).number_format = PCT_FORMAT; col += 1

        indexed_cell = ws.cell(row=ri, column=col, value=rec.get("final_after_indexing_pct", 0) / 100)
        indexed_cell.number_format = PCT_FORMAT
        indexed_cell.font = Font(bold=True)

    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16


def _write_level5_sheet(ws, level_data: dict):
    records = level_data["data"]
    headers = [
        "City", "Hub", "Sub Category", "Cut Class",
        "Baseline", "Hub Drop %", "Drop With Current %",
        "Target SubCat-Cut %", "Final After Indexing %", "Final Rev",
    ]

    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, 1, 1, len(headers))

    for ri, rec in enumerate(records, 2):
        ws.cell(row=ri, column=1, value=rec.get("city_name", ""))
        ws.cell(row=ri, column=2, value=rec.get("hub_name", ""))
        ws.cell(row=ri, column=3, value=rec.get("sub_category", ""))
        ws.cell(row=ri, column=4, value=rec.get("cut_class", ""))
        ws.cell(row=ri, column=5, value=rec.get("baseline", 0)).number_format = NUM_FORMAT
        ws.cell(row=ri, column=6, value=rec.get("hub_drop_pct", 0) / 100).number_format = PCT_FORMAT
        ws.cell(row=ri, column=7, value=rec.get("drop_with_current_pct", 0) / 100).number_format = PCT_FORMAT
        ws.cell(row=ri, column=8, value=rec.get("target_subcat_cut_drop_pct", 0) / 100).number_format = PCT_FORMAT
        cell9 = ws.cell(row=ri, column=9, value=rec.get("final_after_indexing_pct", 0) / 100)
        cell9.number_format = PCT_FORMAT
        cell9.font = Font(bold=True)
        ws.cell(row=ri, column=10, value=rec.get("final_rev", 0)).number_format = NUM_FORMAT

    for c in range(1, 11):
        ws.column_dimensions[get_column_letter(c)].width = 18


def export_all_levels(
    name: str,
    city_data: dict,
    subcat_data: dict,
    subcat_cut_data: dict,
    hub_data: dict,
    hub_cut_data: dict,
) -> bytes:
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "City"
    _write_city_sheet(ws1, city_data)

    ws2 = wb.create_sheet("City-Subcategory")
    _write_indexed_sheet(ws2, subcat_data, ["city_name", "sub_category"], "City Drop %")

    ws3 = wb.create_sheet("City-Subcategory-CutClass")
    _write_indexed_sheet(
        ws3, subcat_cut_data,
        ["city_name", "sub_category", "cut_class"],
        "SubCat Drop %",
    )

    ws4 = wb.create_sheet("City-Hub")
    _write_indexed_sheet(ws4, hub_data, ["city_name", "hub_name"], "City Drop %")

    ws5 = wb.create_sheet("City-Hub-CutClass")
    _write_level5_sheet(ws5, hub_cut_data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
